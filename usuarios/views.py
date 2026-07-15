from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from .forms import RegistroPacienteForm, CitaForm, CitaSecretariaForm, ControlPrenatalForm, EditarPacienteSecretariaForm
from .models import Usuario, LogAuditoria
from citas.models import Cita
from datetime import time, timedelta
from django.db.models import Q, Count
from control_prenatal.models import ControlPrenatal
from django.http import JsonResponse, HttpResponse
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.contrib.auth import update_session_auth_hash
from django.urls import reverse
from django.core.paginator import Paginator
import csv



# ── HELPERS AUDITORÍA ──────────────────────────────────────────

def get_limite_cancelacion():
    """Retorna (hoy, hora_limite) con 30 min de margen para auto-cancelar."""
    ahora_dt = timezone.localtime()
    limite_dt = ahora_dt - timedelta(minutes=30)
    return limite_dt.date(), limite_dt.time()


def auto_cancelar_citas(queryset):
    """Cancela citas pendientes cuya fecha+hora ya superó los 30 min de margen."""
    hoy, limite = get_limite_cancelacion()
    queryset.filter(estado='pendiente').filter(
        Q(fecha__lt=hoy) |
        Q(fecha=hoy, hora__lt=limite)
    ).update(estado='cancelada')


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def registrar_log(request, accion, modulo='', descripcion='', severidad='INFO'):
    """Registra una acción en el log de auditoría."""
    try:
        LogAuditoria.objects.create(
            usuario     = request.user if request.user.is_authenticated else None,
            accion      = accion,
            modulo      = modulo,
            descripcion = descripcion,
            ip_address  = get_client_ip(request),
            severidad   = severidad,
        )
    except Exception:
        pass  # Nunca romper el flujo por un log fallido


def q_usuarios_con_acceso_prenatal():
    return Q(paciente__estado_embarazo='ACTIVO') | Q(citas_paciente__medico__medico__especialidad__tipo='prenatal')


def q_pacientes_con_acceso_prenatal():
    return Q(estado_embarazo='ACTIVO') | Q(usuario__citas_paciente__medico__medico__especialidad__tipo='prenatal')


def q_citas_con_acceso_prenatal():
    return Q(paciente__paciente__estado_embarazo='ACTIVO') | Q(medico__medico__especialidad__tipo='prenatal')


User = get_user_model()
 
from functools import wraps


def no_cache_view(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        response = view_func(request, *args, **kwargs)
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
    return wrapper


def paciente_prenatal_required(view_func):
    """
    Decorator: solo pacientes con módulo prenatal activo pueden acceder.
    Acepta:
      - tipo_paciente == 'prenatal'  (cuentas prenatales puras)
      - tipo_paciente == 'general' con tiene_prenatal == True (cuentas generales con módulo activado)
    """
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.rol != 'paciente':
            return redireccionar_por_rol(request.user)
        if not request.user.puede_prenatal:
            # Es paciente general sin módulo prenatal → su dashboard general
            return redirect('paciente_general_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def medico_required(view_func):
    """Decorator: solo médicos (cualquier especialidad) pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.rol != 'medico':
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def medico_prenatal_required(view_func):
    """Decorator: solo médicos de especialidad PRENATAL pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.rol != 'medico':
            return redireccionar_por_rol(request.user)
        try:
            es_prenatal = (request.user.medico.especialidad and
                           request.user.medico.especialidad.tipo == 'prenatal')
        except Exception:
            es_prenatal = True
        if not es_prenatal:
            from django.contrib import messages as _msg
            _msg.error(request, 'Esta sección es exclusiva para médicos de especialidades prenatales.')
            return redirect('medico_dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def admin_required(view_func):
    """Decorator: solo administradores pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.rol != 'admin':
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper


def secretaria_required(view_func):
    """Decorator: solo secretarias pueden acceder."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.rol != 'secretaria':
            return redireccionar_por_rol(request.user)
        return view_func(request, *args, **kwargs)
    return wrapper

@no_cache_view
def login_view(request):
    # Si ya está autenticado, redirigir directo a su panel
    if request.user.is_authenticated:
        return redireccionar_por_rol(request.user)

    especialidad_id = request.GET.get('especialidad') or request.POST.get('especialidad_id', '')
    tipo_paciente = request.GET.get('tipo') or request.POST.get('tipo_paciente') or 'general'

    if request.method == 'POST':
        username_or_email = request.POST.get('username', '').strip()
        password = request.POST.get('password')
        tipo_esperado = request.POST.get('tipo_paciente') or 'general'

        # Buscar por username O por email
        user_obj = None
        try:
            user_obj = User.objects.get(username=username_or_email)
        except User.DoesNotExist:
            # Intentar buscar por email
            try:
                user_obj = User.objects.get(email__iexact=username_or_email)
            except User.DoesNotExist:
                pass
            except User.MultipleObjectsReturned:
                # Si hay varios usuarios con el mismo email, tomar el primero
                user_obj = User.objects.filter(email__iexact=username_or_email).first()

        if user_obj:
            # Verificar si la cuenta está desactivada
            if user_obj.check_password(password) and not user_obj.is_active:
                registrar_log(request, 'ERROR', 'Autenticación',
                    f'Intento de acceso a cuenta desactivada: "{user_obj.username}"', 'WARNING')
                messages.error(request, 'CUENTA_DESACTIVADA')
                return render(request, 'login.html', {'especialidad_id': especialidad_id, 'tipo_paciente': tipo_paciente})
            # Autenticar usando el username real del usuario encontrado
            user = authenticate(request, username=user_obj.username, password=password)
        else:
            user = None

        if user is not None:
            login(request, user)
            registrar_log(request, 'LOGIN', 'Autenticación',
                f'Inicio de sesión exitoso — rol: {user.rol}', 'INFO')
            
            # Si es paciente, check tipo esperado
            if user.rol == 'paciente':
                if tipo_esperado == 'prenatal' and not user.puede_prenatal:
                    messages.info(request, 'Tu cuenta no tiene el módulo prenatal activo. Solicita a la secretaria que lo active.')
            
            return redireccionar_por_rol(user)
        else:
            # Detectar si fue cuenta desactivada o credenciales incorrectas
            if user_obj and user_obj.check_password(password) and not user_obj.is_active:
                pass  # ya manejado arriba
            else:
                registrar_log(request, 'ERROR', 'Autenticación',
                    f'Intento de login fallido para usuario: "{username_or_email}"', 'WARNING')
            messages.error(request, 'Usuario/correo o contraseña incorrectos')

    return render(request, 'login.html', {'especialidad_id': especialidad_id, 'tipo_paciente': tipo_paciente})
 
 
@no_cache_view
def logout_view(request):

    rol = getattr(request.user, 'rol', None)

    tipo_paciente = getattr(request.user, 'tipo_paciente', None)

    # Guardar nombre ANTES de logout, porque después request.user es AnonymousUser

    nombre_usuario = ''

    if request.user.is_authenticated:

        nombre_usuario = request.user.get_full_name() or request.user.username

    registrar_log(request, 'LOGOUT', 'Autenticación',

        f'Cierre de sesión — usuario: {nombre_usuario}', 'INFO')

    logout(request)
    return redirect('login')
 
 
def redireccionar_por_rol(user):
    if user.rol == 'admin':
        return redirect('admin_dashboard')
    elif user.rol == 'medico':
        return redirect('medico_dashboard')
    elif user.rol == 'secretaria':
        return redirect('secretaria_dashboard')
    elif user.rol == 'paciente':
        # Si es general pura, va al dashboard general
        # Si tiene módulo prenatal activo (sea general o prenatal), va al dashboard prenatal
        if user.puede_prenatal:
            return redirect('paciente_dashboard')
        else:
            return redirect('paciente_general_dashboard')
    else:
        return redirect('login')
 
 
@login_required
@no_cache_view
def admin_dashboard(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from django.utils import timezone

    hoy = timezone.localdate()
    ahora = timezone.localtime().time()

    # ── AUTO-CANCELAR citas pendientes vencidas (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(estado='pendiente'))

    total_usuarios = User.objects.count()
    total_medicos = User.objects.filter(rol='medico').count()
    total_pacientes = Paciente.objects.count()
    total_pacientes_generales = User.objects.filter(rol='paciente', paciente__estado_embarazo='NINGUNO').count()
    total_pacientes_prenatales = User.objects.filter(
        rol='paciente'
    ).filter(q_usuarios_con_acceso_prenatal()).distinct().count()
    total_citas = Cita.objects.count()
    citas_hoy = Cita.objects.filter(fecha=hoy).count()
    citas_pendientes = Cita.objects.filter(estado='pendiente').count()
    citas_atendidas = Cita.objects.filter(estado='atendido').count()
    citas_canceladas = Cita.objects.filter(estado='cancelado').count()
    total_controles = ControlPrenatal.objects.count()
    citas_recientes = Cita.objects.select_related('paciente', 'medico').order_by('-fecha', 'hora')[:8]

    return render(request, 'admin/dashboard_admin.html', {
        'total_usuarios': total_usuarios,
        'total_medicos': total_medicos,
        'total_pacientes': total_pacientes,
        'total_pacientes_generales': total_pacientes_generales,
        'total_pacientes_prenatales': total_pacientes_prenatales,
        'total_citas': total_citas,
        'citas_hoy': citas_hoy,
        'citas_pendientes': citas_pendientes,
        'citas_atendidas': citas_atendidas,
        'citas_canceladas': citas_canceladas,
        'total_controles': total_controles,
        'citas_recientes': citas_recientes,
        'hoy': hoy,
    })
 

@login_required
@no_cache_view
def medico_dashboard(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from citas.models import Cita
    from control_prenatal.models import ControlPrenatal
    from django.db.models import Q

    # Determinar si es médico prenatal o general
    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(medico=request.user))

    # ── STATS ──
    hoy = timezone.now().date()
    total_citas               = Cita.objects.filter(medico=request.user).count()
    citas_pendientes          = Cita.objects.filter(medico=request.user, estado='pendiente', fecha=hoy).count()
    citas_pendientes_futuras  = Cita.objects.filter(medico=request.user, estado='pendiente').count()
    citas_atendidas           = Cita.objects.filter(medico=request.user, estado='atendido').count()
    citas_canceladas          = Cita.objects.filter(medico=request.user, estado__in=['cancelado', 'cancelada']).count()
    proximas_citas            = Cita.objects.filter(medico=request.user, estado='pendiente').order_by('fecha', 'hora')[:5]

    # ── PACIENTES: solo del tipo que corresponde al médico ──
    if es_prenatal:
        total_pacientes = Usuario.objects.filter(
            rol='paciente'
        ).filter(q_usuarios_con_acceso_prenatal()).distinct().count()
    else:
        total_pacientes = Usuario.objects.filter(
            rol='paciente',
            paciente__estado_embarazo='NINGUNO'
        ).count()

    context = {
        'total_citas':              total_citas,
        'citas_pendientes':         citas_pendientes,
        'citas_pendientes_futuras': citas_pendientes_futuras,
        'citas_atendidas':          citas_atendidas,
        'citas_canceladas':         citas_canceladas,
        'proximas_citas':           proximas_citas,
        'total_pacientes':          total_pacientes,
        'es_prenatal':              es_prenatal,
    }

    if es_prenatal:
        total_controles = ControlPrenatal.objects.filter(medico=request.user).count()
        context['total_controles'] = total_controles
        return render(request, 'medico/dashboard_medico.html', context)
    else:
        return render(request, 'medico/dashboard_medico_general.html', context)
 
 
@login_required
@no_cache_view
def secretaria_dashboard(request):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects)
    hoy = timezone.now().date()

    from pacientes.models import Paciente
    return render(request, 'secretaria/dashboard_secretaria.html', {
        'total_citas':      Cita.objects.count(),
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
        'citas_atendidas':  Cita.objects.filter(estado='atendido').count(),
        'total_pacientes':  Paciente.objects.count(),
        'citas_recientes':  Cita.objects.order_by('-fecha', 'hora')[:5],
    })
 
 
@paciente_prenatal_required
@no_cache_view
def paciente_dashboard(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(paciente=request.user))
    hoy = timezone.now().date()

    from pacientes.models import Paciente
    from paciente_general.models import ProgramacionParto
    try:
        perfil = Paciente.objects.get(usuario=request.user)
    except Paciente.DoesNotExist:
        perfil = None

    # ── Programación de parto próxima ─────────────────────────────────────
    parto_programado = ProgramacionParto.objects.filter(
        paciente=request.user,
        estado__in=['programado', 'confirmado'],
        fecha_programada__gte=hoy,
    ).order_by('fecha_programada').first()

    response = render(request, 'paciente/dashboard_paciente.html', {
        'perfil': perfil,
        'parto_programado': parto_programado,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    return response
 
def registro_paciente(request):
    # ?tipo=ginecologia, ?tipo=medicina, etc. (podemos recibir la especialidad)
    especialidad_param = request.GET.get('tipo') or request.POST.get('especialidad_param') or ''

    if request.method == 'POST':
        form = RegistroPacienteForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.rol = 'paciente'
            user.genero = form.cleaned_data.get('genero', '')
            user.save()

            from pacientes.models import Paciente
            paciente, _ = Paciente.objects.get_or_create(usuario=user)
            paciente.cedula = form.cleaned_data.get('cedula', '')
            paciente.telefono = form.cleaned_data.get('telefono', '')
            paciente.estado_embarazo = 'NINGUNO'
            paciente.save()

            login(request, user)
            messages.success(request, f'¡Cuenta creada con éxito! Bienvenido/a, {user.first_name}.')
            return redirect('paciente_general_dashboard')
        else:
            # Recopilar todos los errores del form en mensajes legibles
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        label = form.fields[field].label if field in form.fields else field
                        messages.error(request, f'{label}: {error}')
    else:
        form = RegistroPacienteForm()

    return render(request, 'registro.html', {
        'form': form,
        'especialidad_param': especialidad_param,
    })


def verificar_disponibilidad(request):
    """
    Endpoint AJAX (GET) — verifica si username o cédula ya existen.
    ?campo=username&valor=XXX  |  ?campo=cedula&valor=XXX
    """
    campo = request.GET.get('campo', '').strip()
    valor = request.GET.get('valor', '').strip()

    if not campo or not valor:
        return JsonResponse({'disponible': False, 'mensaje': 'Datos incompletos.'})

    if campo == 'username':
        existe = Usuario.objects.filter(username__iexact=valor).exists()
        if existe:
            return JsonResponse({'disponible': False, 'mensaje': 'Este usuario ya está registrado.'})
        return JsonResponse({'disponible': True, 'mensaje': 'Usuario disponible.'})

    if campo == 'cedula':
        from pacientes.models import Paciente
        existe = Paciente.objects.filter(cedula=valor).exists()
        if existe:
            return JsonResponse({'disponible': False, 'mensaje': 'Esta cédula ya está registrada.'})
        return JsonResponse({'disponible': True, 'mensaje': 'Cédula disponible.'})

    return JsonResponse({'disponible': False, 'mensaje': 'Campo no válido.'})


@paciente_prenatal_required
@no_cache_view
def mi_perfil(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'perfil':
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name = request.POST.get('last_name', '')
            request.user.email = request.POST.get('email', '')
            request.user.save()
            messages.success(request, 'Datos actualizados correctamente.')

        elif action == 'password':
            from django.contrib.auth import update_session_auth_hash
            password_actual = request.POST.get('password_actual')
            password_nuevo = request.POST.get('password_nuevo')
            password_confirmar = request.POST.get('password_confirmar')

            if not request.user.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif password_nuevo != password_confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(password_nuevo) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(password_nuevo)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña actualizada correctamente.')

        return redirect('mi_perfil')

    return render(request, 'paciente/mi_perfil.html')
 
 
@paciente_prenatal_required
@no_cache_view
def agendar_cita(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    from medicos.models import Medico
    from landing.models import Especialidad

    # Todas las especialidades (excepto partos) y todos los médicos activos
    especialidades_prenatales = Especialidad.objects.filter(activo=True).exclude(nombre__icontains='parto').exclude(nombre__icontains='cesárea')
    medicos_prenatales = Medico.objects.filter(
        usuario__is_active=True
    ).select_related('usuario', 'especialidad')

    if request.method == 'POST':
        medico_id  = request.POST.get('medico_id')
        fecha      = request.POST.get('fecha')
        hora       = request.POST.get('hora')
        motivo     = request.POST.get('motivo', '').strip()

        errores = []
        if not medico_id:
            errores.append('Debes seleccionar un médico.')
        if not fecha:
            errores.append('Debes seleccionar una fecha.')
        if not hora:
            errores.append('Debes seleccionar una hora.')

        if not errores:
            try:
                medico_usuario = Usuario.objects.get(id=medico_id, rol='medico')
                existe = Cita.objects.filter(
                    medico=medico_usuario, fecha=fecha, hora=hora,
                    estado__in=['pendiente', 'confirmada']
                ).exists()
                if existe:
                    errores.append('Esta hora ya está ocupada para el médico seleccionado. Elige otra hora.')
                else:
                    cita = Cita(
                        paciente=request.user,
                        medico=medico_usuario,
                        fecha=fecha,
                        hora=hora,
                        motivo=motivo,
                    )
                    # Asignar especialidad del médico automáticamente
                    try:
                        perfil_medico = Medico.objects.get(usuario=medico_usuario)
                        if perfil_medico.especialidad:
                            cita.especialidad = perfil_medico.especialidad
                    except Medico.DoesNotExist:
                        pass
                    cita.save()
                    registrar_log(request, 'CREATE', 'Citas',
                        f'Cita agendada por {request.user.get_full_name()} '
                        f'para el {fecha} a las {hora}')
                    messages.success(request, 'Cita agendada correctamente.')
                    return redirect('ver_citas')
            except Usuario.DoesNotExist:
                errores.append('El médico seleccionado no es válido.')

        for e in errores:
            messages.error(request, e)

    # Horas disponibles para precarga (todas)
    horas = [
        '08:30', '09:00', '09:30', '10:00', '10:30', '11:00', '11:30',
        '12:00', '12:30', '14:00', '14:30', '15:00', '15:30', '16:00', '16:30',
    ]

    return render(request, 'paciente/agendar_cita.html', {
        'medicos': medicos_prenatales,
        'especialidades': especialidades_prenatales,
        'horas': horas,
    })
 
@paciente_prenatal_required
@no_cache_view
def ver_citas(request):
    # paciente_prenatal_required ya valida rol y tipo_paciente

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(paciente=request.user))
    hoy = timezone.now().date()

    citas = Cita.objects.filter(paciente=request.user).order_by('-fecha')

    return render(request, 'paciente/ver_citas.html', {'citas': citas})
 
@login_required
@no_cache_view
def citas_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects.filter(medico=request.user))
    hoy = timezone.now().date()

    citas = Cita.objects.filter(medico=request.user).order_by('fecha', 'hora')

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    template = 'medico/citas_medico.html' if es_prenatal else 'medico/citas_medico_general.html'
    return render(request, template, {'citas': citas})
 
@login_required
def cambiar_estado_cita(request, cita_id):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)
 
    cita = get_object_or_404(Cita, id=cita_id, medico=request.user)
 
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in ['pendiente', 'confirmada', 'atendido', 'cancelado']:
            cita.estado = nuevo_estado
            cita.save()
            registrar_log(request, 'UPDATE', 'Citas', f"Cambió estado de la cita {cita.id} a {nuevo_estado}", 'INFO')
            labels = {'confirmada': 'aceptada', 'cancelado': 'rechazada', 'atendido': 'marcada como atendida', 'pendiente': 'vuelta a pendiente'}
            messages.success(request, f'Cita {labels.get(nuevo_estado, "actualizada")} correctamente.')

    return redirect('citas_medico')
 
@login_required
@no_cache_view
def pacientes_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    from citas.models import Cita

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except Exception:
        es_prenatal = True

    if es_prenatal:
        # ── Médico prenatal: ve sus pacientes asignadas + pacientes con citas
        #    pendientes/atendidas con él que aún no tienen expediente prenatal ──
        #
        # 1) Pacientes que ya tienen medico_prenatal = este médico
        ids_asignadas = Paciente.objects.filter(
            medico_prenatal=request.user
        ).values_list('usuario_id', flat=True)

        # 2) Pacientes con citas con este médico (para poder activar embarazo)
        ids_con_citas = Cita.objects.filter(
            medico=request.user
        ).values_list('paciente_id', flat=True)

        # Unión de ambos conjuntos
        todos_ids = set(ids_asignadas) | set(ids_con_citas)

        pacientes = Paciente.objects.filter(
            usuario_id__in=todos_ids
        ).select_related('usuario', 'medico_prenatal').order_by(
            'usuario__first_name', 'usuario__last_name'
        )

        return render(request, 'medico/pacientes.html', {'pacientes': pacientes})

    else:
        # ── Médico general: ve solo pacientes con citas registradas con él ──
        ids_con_citas = Cita.objects.filter(
            medico=request.user
        ).values_list('paciente_id', flat=True)

        pacientes = Paciente.objects.filter(
            usuario_id__in=ids_con_citas
        ).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name')

        pacientes_activos = pacientes.filter(usuario__is_active=True).count()

        return render(request, 'medico/pacientes_general.html', {
            'pacientes': pacientes,
            'pacientes_activos': pacientes_activos,
            'total': pacientes.count(),
        })

@login_required
def activar_embarazo(request, paciente_id):
    if request.user.rol not in ['medico', 'secretaria', 'admin']:
        return redirect('landing')
    
    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
    paciente.estado_embarazo = 'ACTIVO'

    # Si quien activa es un médico prenatal, se asigna como médico responsable
    if request.user.rol == 'medico':
        try:
            es_prenatal = request.user.medico.especialidad and \
                          request.user.medico.especialidad.tipo == 'prenatal'
        except Exception:
            es_prenatal = False
        if es_prenatal and not paciente.medico_prenatal:
            paciente.medico_prenatal = request.user

    paciente.save()
    
    messages.success(request, f'Expediente prenatal activado para {paciente.usuario.get_full_name()}.')
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('pacientes_medico')


@login_required
def desactivar_embarazo(request, paciente_id):
    if request.user.rol not in ['medico', 'secretaria', 'admin']:
        return redirect('landing')

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
    nombre = paciente.usuario.get_full_name()
    paciente.estado_embarazo = 'FINALIZADO'
    paciente.mensaje_prenatal_visto = False
    paciente.save()

    registrar_log(request, 'UPDATE', 'Pacientes',
        f'Embarazo desactivado para {nombre} por {request.user.get_full_name()}', 'INFO')
    messages.success(request, f'Módulo prenatal desactivado para {nombre}. La paciente volvió a modo general.')
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    return redirect('pacientes_medico')
@medico_prenatal_required
@login_required
def registrar_control(request):
    """
    Solo médicos de especialidad PRENATAL pueden registrar controles.
    Al guardar, la IA se ejecuta AUTOMÁTICAMENTE y guarda la predicción.
    """
    # medico_prenatal_required ya validó rol y especialidad

    if request.method == 'POST':
        form = ControlPrenatalForm(request.POST)
        if form.is_valid():
            control = form.save(commit=False)
            control.medico = request.user
            control.save()
            registrar_log(request, 'CREATE', 'Controles Prenatales', f"Se registró control para paciente ID {control.paciente.id}", 'INFO')

            # ── Disparar IA automáticamente ────────────────────────────────
            prediccion = _ejecutar_ia_en_control(control, request)

            if prediccion:
                messages.success(request, f'Control registrado. La IA determinó riesgo {prediccion.nivel_riesgo} ({prediccion.puntuacion_riesgo}%) — revisa el panel clínico completo.')
                return redirect('ver_predicciones_paciente', paciente_id=control.paciente.id)
            else:
                messages.success(request, 'Control prenatal registrado. No se pudo generar la evaluación IA automáticamente.')
                return redirect('historial_prenatal')
    else:
        form = ControlPrenatalForm()

    return render(request, 'medico/registrar_control.html', {'form': form})


def _ejecutar_ia_en_control(control, request=None):
    """
    Ejecuta el modelo ML con los datos del control prenatal y guarda PrediccionIA.
    Devuelve el objeto PrediccionIA creado o None si el modelo no está disponible.

    Unidades: glucosa en mg/dL (se convierte internamente a mmol/L para el modelo).
    """
    try:
        from ai.predict import engine as ml_engine
        from prediccion_ia.models import PrediccionIA
        from pacientes.models import Paciente
        import json
        import logging
        logger = logging.getLogger(__name__)

        if not ml_engine.modelo_disponible:
            logger.warning('[IA] Modelo no disponible — saltando prediccion')
            return None

        # Obtener perfil del paciente
        try:
            perfil = Paciente.objects.get(usuario=control.paciente)
        except Paciente.DoesNotExist:
            logger.error(f'[IA] No existe perfil Paciente para usuario {control.paciente}')
            return None

        # Edad del paciente
        edad = int(perfil.edad or 25)
        if not (10 <= edad <= 60):
            edad = 25

        # Parsear presion arterial con validacion
        try:
            sistolica  = int(control.presion_sistolica)
            diastolica = int(control.presion_diastolica)
            if not (60 <= sistolica <= 200):
                sistolica = 120
            if not (40 <= diastolica <= 140):
                diastolica = 80
        except Exception:
            sistolica, diastolica = 120, 80

        # BMI calculado desde peso/altura (mas preciso)
        try:
            altura = float(control.altura)
            peso   = float(control.peso)
            if altura > 0 and peso > 0:
                bmi = round(peso / (altura ** 2), 2)
            else:
                bmi = 23.0
            bmi = max(10.0, min(bmi, 60.0))
        except Exception:
            bmi = 23.0

        # Glucosa: mg/dL -> mmol/L para el modelo
        try:
            glucosa_mgdl = float(control.glucosa)
            if glucosa_mgdl <= 0:
                glucosa_mgdl = 90.0
            glucosa_mmol = round(glucosa_mgdl / 18.0, 3)
            glucosa_mmol = max(3.0, min(glucosa_mmol, 25.0))
        except Exception:
            glucosa_mmol = 5.0
            glucosa_mgdl = 90.0

        # Frecuencia cardiaca y temperatura con validacion
        try:
            fc = int(control.frecuencia_cardiaca)
            fc = max(40, min(fc, 180))
        except Exception:
            fc = 75

        try:
            temp = float(control.temperatura)
            temp = max(95.0, min(temp, 106.0))
        except Exception:
            temp = 98.0

        # Vector de datos clinicos para el modelo
        datos_clinicos = {
            'age':                  edad,
            'systolic_bp':          sistolica,
            'diastolic_bp':         diastolica,
            'glucose':              glucosa_mmol,
            'body_temp':            temp,
            'heart_rate':           fc,
            'bmi':                  bmi,
            'prev_complications':   int(bool(control.complicaciones_previas)),
            'diabetes_preexisting': int(bool(control.diabetes_preexistente)),
            'diabetes_gestacional': int(bool(control.diabetes_gestacional)),
        }

        logger.info(f'[IA] Prediccion para paciente {perfil} -- datos: {datos_clinicos}')
        resultado = ml_engine.predict(datos_clinicos)
        logger.info(f'[IA] Resultado: {resultado.nivel_riesgo} ({resultado.puntuacion}%)')

        resultado_json = json.dumps({
            'factores':        resultado.factores_detectados,
            'complicaciones':  resultado.complicaciones,
            'recomendaciones': resultado.recomendaciones,
            'recomendaciones_grupos': resultado.recomendaciones_grupos,
            'probabilidades':  resultado.probabilidades,
            'explicacion':     resultado.explicacion,
            'alerta_critica':  resultado.alerta_critica,
            'nota_antecedente': resultado.nota_antecedente,
            'datos_procesados': {
                'glucosa_mmol':  glucosa_mmol,
                'glucosa_mgdl':  round(float(control.glucosa), 1),
                'bmi_calculado': bmi,
                'sistolica':     sistolica,
                'diastolica':    diastolica,
            }
        }, ensure_ascii=False)

        prediccion = PrediccionIA.objects.create(
            paciente               = perfil,
            edad                   = edad,
            semanas_gestacion      = int(control.semanas_gestacion),
            presion_arterial       = str(control.presion_arterial),
            presion_sistolica      = sistolica,
            presion_diastolica     = diastolica,
            peso                   = float(control.peso),
            altura                 = float(control.altura),
            imc                    = bmi,
            glucosa                = float(control.glucosa),
            frecuencia_cardiaca    = fc,
            temperatura            = temp,
            embarazos_previos      = int(control.embarazos_previos),
            complicaciones_previas = bool(control.complicaciones_previas),
            diabetes_preexistente  = bool(control.diabetes_preexistente),
            diabetes_gestacional   = bool(control.diabetes_gestacional),
            nivel_riesgo           = resultado.nivel_riesgo,
            puntuacion_riesgo      = resultado.puntuacion,
            resultado              = resultado_json,
        )
        return prediccion

    except Exception as e:
        import logging, traceback
        logging.getLogger(__name__).error(
            f'[IA] Error al ejecutar prediccion: {e}\n{traceback.format_exc()}'
        )
        return None

@medico_prenatal_required
@login_required
def historial_prenatal(request):
    # medico_prenatal_required ya validó rol y especialidad
 
    paciente_id = request.GET.get('paciente')

    if paciente_id:
        controles = ControlPrenatal.objects.filter(
            paciente_id=paciente_id
        ).select_related('paciente', 'medico').order_by('-fecha')
    else:
        controles = ControlPrenatal.objects.select_related(
            'paciente', 'medico'
        ).order_by('-fecha')

    # ── Enriquecer cada control con su predicción IA ──────────────────────
    from prediccion_ia.models import PrediccionIA
    from pacientes.models import Paciente
    import json

    # Obtener IDs de usuarios de los controles para una sola consulta
    paciente_ids = list({c.paciente_id for c in controles})
    perfiles_map = {
        p.usuario_id: p
        for p in Paciente.objects.filter(usuario_id__in=paciente_ids)
    }

    # Obtener la última predicción de cada paciente en un solo query
    predicciones_map = {}
    for pred in PrediccionIA.objects.filter(
        paciente__usuario_id__in=paciente_ids
    ).select_related('paciente').order_by('paciente', '-fecha'):
        pid = pred.paciente.usuario_id
        if pid not in predicciones_map:
            predicciones_map[pid] = pred

    # Adjuntar al control su predicción más reciente
    controles_enriquecidos = []
    for c in controles:
        c.prediccion_ia = predicciones_map.get(c.paciente_id)
        c.imc_calculado = c.imc
        controles_enriquecidos.append(c)

    pacientes = Usuario.objects.filter(
        rol='paciente'
    ).filter(q_usuarios_con_acceso_prenatal()).distinct().order_by('first_name', 'last_name')
 
    return render(request, 'medico/historial.html', {
        'controles': controles_enriquecidos,
        'pacientes': pacientes,
        'paciente_filtrado_id': int(paciente_id) if paciente_id else None,
    })
 
@login_required
def registrar_paciente(request, tipo='prenatal'):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)

    # Validar tipo
    if tipo not in ('prenatal', 'general'):
        tipo = 'prenatal'

    if request.method == 'POST':
        form = RegistroPacienteForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')

            # Verifica si el usuario ya existe
            if User.objects.filter(username=username).exists():
                messages.error(request, f'Ya existe una paciente con el usuario "{username}". Por favor elige otro.')
                return render(request, 'secretaria/registrar_paciente.html', {'form': form, 'tipo': tipo})

            user = form.save(commit=False)
            user.rol = 'paciente'
            user.paciente.estado_embarazo = 'ACTIVO' if tipo == 'prenatal' else 'NINGUNO'
            user.paciente.save()
            user.genero = form.cleaned_data.get('genero', '')
            user.save()

            from pacientes.models import Paciente
            paciente, _ = Paciente.objects.get_or_create(usuario=user)
            paciente.cedula = form.cleaned_data.get('cedula', '')
            paciente.telefono = form.cleaned_data.get('telefono', '')
            paciente.save()

            messages.success(request, f'Paciente {user.get_full_name()} registrada correctamente como paciente {tipo}.')
            return redirect('lista_pacientes_secretaria')
    else:
        form = RegistroPacienteForm()

    return render(request, 'secretaria/registrar_paciente.html', {'form': form, 'tipo': tipo})


@login_required
@no_cache_view
def lista_pacientes_secretaria(request):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    # Prenatales = tipo prenatal O general con módulo prenatal activado
    pacientes_prenatales = Paciente.objects.filter(
        q_pacientes_con_acceso_prenatal()
    ).select_related('usuario').distinct()
    pacientes_generales = Paciente.objects.exclude(
        q_pacientes_con_acceso_prenatal()
    ).select_related('usuario').distinct()

    return render(request, 'secretaria/lista_pacientes_secretaria.html', {
        'pacientes_prenatales': pacientes_prenatales,
        'pacientes_generales':  pacientes_generales,
    })


@login_required
def toggle_modulo_prenatal(request, paciente_id):
    """
    La secretaria activa o desactiva el módulo prenatal para una paciente general.
    Solo POST. Solo accesible por secretaria o admin.
    """
    from pacientes.models import Paciente

    if request.user.rol not in ('secretaria', 'admin'):
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')
    perfil, _ = Paciente.objects.get_or_create(usuario=paciente_usuario)

    if request.method == 'POST':
        activar = request.POST.get('activar') == '1'
        if paciente_usuario.paciente.estado_embarazo == 'ACTIVO':
            messages.error(request, 'Solo las pacientes generales usan activación de módulo prenatal.')
            return redirect(request.POST.get('next', 'lista_pacientes_secretaria'))
        if activar and paciente_usuario.genero != 'femenino':
            messages.error(request, 'Solo las pacientes registradas como femeninas pueden activar el módulo prenatal.')
            return redirect(request.POST.get('next', 'lista_pacientes_secretaria'))
        perfil.tiene_prenatal = activar
        perfil.save()

        if activar:
            messages.success(
                request,
                f'Módulo prenatal ACTIVADO para {paciente_usuario.get_full_name() or paciente_usuario.username}. '
                f'Ahora puede acceder a controles prenatales, IA y chatbot.'
            )
        else:
            messages.success(
                request,
                f'Módulo prenatal desactivado para {paciente_usuario.get_full_name() or paciente_usuario.username}.'
            )

    return redirect(request.POST.get('next', 'lista_pacientes_secretaria'))

 
@login_required
def crear_historia_clinica(request, paciente_id):
    """Crea o edita la Historia Clínica Obstétrica de una paciente."""
    from control_prenatal.models import HistoriaClinica
    from usuarios.forms import HistoriaClinicaForm

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')

    # Si ya existe, redirigir a editar
    historia_existente = HistoriaClinica.objects.filter(paciente=paciente_usuario).first()
    if historia_existente:
        return redirect('editar_historia_clinica', historia_id=historia_existente.id)

    if request.method == 'POST':
        form = HistoriaClinicaForm(request.POST)
        if form.is_valid():
            historia = form.save(commit=False)
            historia.medico = request.user
            historia.save()
            registrar_log(request, 'CREATE', 'Historia Clínica', f"Historia clínica creada para paciente ID {paciente_usuario.id}", 'INFO')
            messages.success(request, f'Historia clínica de {paciente_usuario.get_full_name()} creada correctamente.')
            return redirect('ver_historia_clinica', paciente_id=paciente_id)
    else:
        form = HistoriaClinicaForm(initial={'paciente': paciente_usuario})
        # Bloquear el campo paciente ya que viene predefinido
        form.fields['paciente'].widget.attrs['disabled'] = True

    return render(request, 'medico/historia_clinica_form.html', {
        'form': form,
        'paciente': paciente_usuario,
        'modo': 'crear',
    })


@login_required
def editar_historia_clinica(request, historia_id):
    """Edita una Historia Clínica Obstétrica existente."""
    from control_prenatal.models import HistoriaClinica
    from usuarios.forms import HistoriaClinicaForm

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    historia = get_object_or_404(HistoriaClinica, id=historia_id)

    if request.method == 'POST':
        form = HistoriaClinicaForm(request.POST, instance=historia)
        if form.is_valid():
            form.save()
            registrar_log(request, 'UPDATE', 'Historia Clínica', f"Historia clínica editada ID {historia.id}", 'INFO')
            messages.success(request, 'Historia clínica actualizada correctamente.')
            return redirect('ver_historia_clinica', paciente_id=historia.paciente_id)
    else:
        form = HistoriaClinicaForm(instance=historia)

    return render(request, 'medico/historia_clinica_form.html', {
        'form': form,
        'paciente': historia.paciente,
        'historia': historia,
        'modo': 'editar',
    })


@login_required
def ver_historia_clinica(request, paciente_id):
    """
    Dashboard completo de la Historia Clínica Obstétrica.
    Muestra: Historia inicial + todos los controles + gráfico de ganancia de peso.
    """
    from control_prenatal.models import HistoriaClinica, ControlPrenatal
    from prediccion_ia.models import PrediccionIA
    from pacientes.models import Paciente
    import json

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    paciente_usuario = get_object_or_404(Usuario, id=paciente_id, rol='paciente')
    perfil = Paciente.objects.filter(usuario=paciente_usuario).first()
    historia = HistoriaClinica.objects.filter(paciente=paciente_usuario).first()

    # Todos los controles ordenados por fecha
    controles = ControlPrenatal.objects.filter(
        paciente=paciente_usuario
    ).order_by('fecha')

    # Última predicción IA
    ultima_prediccion = None
    if perfil:
        ultima_prediccion = PrediccionIA.objects.filter(paciente=perfil).order_by('-fecha').first()

    # ── Datos para el gráfico de ganancia de peso ──────────────────────────
    # Lista de puntos: [{semana, peso, ganancia_desde_inicio}]
    peso_inicial = (historia.peso_inicial if historia and historia.peso_inicial
                    else (controles.first().peso if controles.exists() else None))

    datos_peso_chart = []
    for c in controles:
        ganancia = round(c.peso - peso_inicial, 1) if peso_inicial else 0
        datos_peso_chart.append({
            'semana': c.semanas_gestacion,
            'peso': c.peso,
            'ganancia': ganancia,
            'fecha': c.fecha.strftime('%d/%m/%Y'),
        })

    # Rangos recomendados de ganancia para las bandas del gráfico
    rango_ganancia = None
    if historia:
        rango_ganancia = historia.ganancia_recomendada

    return render(request, 'medico/historia_clinica.html', {
        'paciente':          paciente_usuario,
        'perfil':            perfil,
        'historia':          historia,
        'controles':         controles,
        'ultima_prediccion': ultima_prediccion,
        'datos_peso_json':   json.dumps(datos_peso_chart),
        'rango_ganancia':    rango_ganancia,
        'peso_inicial':      peso_inicial,
    })


@login_required
def editar_perfil_paciente(request, paciente_id):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)
 
    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)
 
    if request.method == 'POST':
        paciente.cedula = request.POST.get('cedula', '')
        paciente.edad = request.POST.get('edad')
        paciente.direccion = request.POST.get('direccion', '')
        paciente.telefono = request.POST.get('telefono', '')
        paciente.fecha_ultima_menstruacion = request.POST.get('fecha_ultima_menstruacion') or None
        paciente.fecha_probable_parto = request.POST.get('fecha_probable_parto') or None
        paciente.save()
        messages.success(request, f'Datos clínicos de {paciente.usuario.get_full_name()} actualizados correctamente.')
        return redirect('pacientes_medico')
 
    return render(request, 'medico/editar_paciente.html', {'paciente': paciente})
 
#HORARIOS DEFINIDOS
HORAS_DISPONIBLES = [
    time(8,30), time(9,0), time(9,30), time(10,0),
    time(10,30), time(11,0), time(11,30),
    time(12,0), time(12,30),
    time(14,0), time(14,30), time(15,0),
    time(15,30), time(16,0), time(16,30)
]
 
# AGENDAR CITA (SECRETARIA)
@login_required
def agendar_cita_secretaria(request):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)

    form = CitaSecretariaForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            cita = form.save(commit=False)
            existe = Cita.objects.filter(
                medico=cita.medico,
                fecha=cita.fecha,
                hora=cita.hora
            ).exists()
            if existe:
                form.add_error('hora', 'Esta hora ya está ocupada.')
            else:
                # Asignar especialidad del médico automáticamente
                try:
                    from medicos.models import Medico
                    medico_perfil = Medico.objects.get(usuario=cita.medico)
                    if medico_perfil.especialidad:
                        cita.especialidad = medico_perfil.especialidad
                except Exception:
                    pass
                cita.save()
                messages.success(request, 'Cita agendada correctamente.')
                registrar_log(request, 'CREATE', 'Citas',
                    f'Cita agendada para {cita.paciente.get_full_name()} el {cita.fecha} a las {cita.hora}', 'INFO')
                return redirect('citas_secretaria')

    # Pasar médicos con sus especialidades para el JS del template
    from medicos.models import Medico
    medicos_qs = Medico.objects.select_related('usuario', 'especialidad').filter(usuario__is_active=True)
    medicos_data = [
        {
            'usuario_id': m.usuario.id,
            'especialidad__nombre': m.especialidad.nombre if m.especialidad else None,
        }
        for m in medicos_qs
    ]

    return render(request, 'secretaria/agendar_cita.html', {
        'form': form,
        'medicos_especialidades': medicos_data,
    })
 
 
# OBTENER HORAS DISPONIBLES
@login_required
def obtener_horas_disponibles(request):
    fecha = request.GET.get('fecha')
    medico = request.GET.get('medico')
 
    if not fecha or not medico:
        return JsonResponse({'horas': []})
 
    horas_ocupadas = Cita.objects.filter(
        fecha=fecha,
        medico_id=medico
    ).values_list('hora', flat=True)
 
    horas_libres = [
        h.strftime("%H:%M")
        for h in HORAS_DISPONIBLES if h not in horas_ocupadas
    ]
 
    return JsonResponse({'horas': horas_libres})
@login_required
def editar_paciente_secretaria(request, paciente_id):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)
    paciente = get_object_or_404(Usuario, id=paciente_id)
    if request.method == 'POST':
        form = EditarPacienteSecretariaForm(request.POST, instance=paciente)
        if form.is_valid():
            # Validar cédula duplicada (excluir el propio paciente)
            from pacientes.models import Paciente
            cedula_nueva = form.cleaned_data.get('cedula', '').strip()
            if cedula_nueva:
                duplicado = Paciente.objects.filter(cedula=cedula_nueva).exclude(usuario=paciente).first()
                if duplicado:
                    messages.error(request, f"¡La cédula {cedula_nueva} ya está registrada para otro paciente ({duplicado.usuario.get_full_name()}).")
                    return render(request, 'secretaria/editar_paciente_secretaria.html', {
                        'form': form, 'paciente': paciente, 'error_cedula': True
                    })
            form.save()
            # Actualizar cédula y teléfono en modelo Paciente
            from pacientes.models import Paciente
            perfil, _ = Paciente.objects.get_or_create(usuario=paciente)
            if cedula_nueva:
                perfil.cedula = cedula_nueva
            telefono_nuevo = form.cleaned_data.get('telefono', '').strip()
            if telefono_nuevo:
                perfil.telefono = telefono_nuevo
            perfil.save()
            registrar_log(request, 'UPDATE', 'Secretaría', f"Se editó el paciente {paciente.get_full_name()} (ID: {paciente.id})", 'INFO')
            messages.success(request, f"Paciente {paciente.get_full_name()} actualizado correctamente.")
            return redirect('lista_pacientes_secretaria')
        else:
            messages.error(request, "Por favor corrige los errores del formulario.")
    else:
        form = EditarPacienteSecretariaForm(instance=paciente)
    
    return render(request, 'secretaria/editar_paciente_secretaria.html', {
        'form': form,
        'paciente': paciente
    })

@login_required
def reprogramar_cita(request, cita_id):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)
 
    cita = Cita.objects.get(id=cita_id)
 
    if request.method == 'POST':
        cita.fecha = request.POST.get('fecha')
        cita.hora = request.POST.get('hora')
        cita.save()
        registrar_log(request, 'UPDATE', 'Citas', f"Secretaria reprogramó cita {cita.id} a {cita.fecha} {cita.hora}", 'INFO')
        return redirect('citas_secretaria')
 
    return render(request, 'secretaria/reprogramar.html', {
        'cita': cita,
        'hoy': timezone.now().date(),
    })
 
@login_required
@no_cache_view
def citas_secretaria(request):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects)
    hoy = timezone.now().date()

    citas_prenatales = Cita.objects.filter(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').distinct().order_by('-fecha', 'hora')

    citas_generales = Cita.objects.exclude(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')

    return render(request, 'secretaria/citas.html', {
        'citas_prenatales': citas_prenatales,
        'citas_generales': citas_generales,
        'total_prenatales': citas_prenatales.count(),
        'total_generales': citas_generales.count(),
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
@login_required
def datos_paciente(request):
    paciente_id = request.GET.get('paciente_id')
    if not paciente_id:
        return JsonResponse({'error': 'No se proporcionó paciente'}, status=400)
 
    try:
        from pacientes.models import Paciente
        paciente = Paciente.objects.get(usuario_id=paciente_id)
        return JsonResponse({
            'cedula': paciente.cedula or '',
            'telefono': paciente.telefono or '',
            'email': paciente.usuario.email or '',
        })
    except Paciente.DoesNotExist:
        return JsonResponse({'cedula': '', 'telefono': '', 'email': ''})
 
@login_required
def buscar_pacientes(request):
    """
    Endpoint AJAX para el buscador en tiempo real de pacientes.
    Busca por nombre, apellido o cédula.
    Retorna: {"pacientes": [{"id": ..., "nombre": ..., "cedula": ..., "email": ...}]}
    """
    if request.user.rol not in ('secretaria', 'admin'):
        return JsonResponse({'pacientes': []}, status=403)
 
    q = request.GET.get('q', '').strip()
    if not q or len(q) < 1:
        return JsonResponse({'pacientes': []})
 
    from pacientes.models import Paciente
    from django.db.models import Q
 
    pacientes = Paciente.objects.filter(
        Q(usuario__first_name__icontains=q) |
        Q(usuario__last_name__icontains=q)  |
        Q(cedula__icontains=q)              |
        Q(usuario__email__icontains=q)
    ).select_related('usuario')[:15]
 
    resultados = []
    for p in pacientes:
        nombre_completo = f"{p.usuario.first_name} {p.usuario.last_name}".strip()
        if not nombre_completo:
            nombre_completo = p.usuario.username
        resultados.append({
            'id':     p.usuario.id,   # mismo ID que usa datos_paciente (usuario_id)
            'nombre': nombre_completo,
            'cedula': p.cedula or '',
            'email':  p.usuario.email or '',
        })
 
    return JsonResponse({'pacientes': resultados})
 
#ADMIN
@login_required
@no_cache_view
def lista_usuarios(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    usuarios = User.objects.all()
 
    # Obtener IDs de usuarios con sesión activa
    sesiones_activas = Session.objects.filter(expire_date__gte=timezone.now())
    ids_activos = set()
    for sesion in sesiones_activas:
        datos = sesion.get_decoded()
        uid = datos.get('_auth_user_id')
        if uid:
            ids_activos.add(int(uid))
 
    # Agregar flag a cada usuario
    for u in usuarios:
        u.sesion_activa = u.id in ids_activos
 
    return render(request, 'admin/lista_usuarios.html', {
        'usuarios': usuarios,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
# ─────────────────────────────────────────────────────────────
# AGREGAR ESTAS DOS VIEWS NUEVAS
# ─────────────────────────────────────────────────────────────
 
@login_required
def toggle_usuario(request, usuario_id):
    """Activa o desactiva la cuenta de un usuario (solo admin, no se puede tocar a otro admin)."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    if request.method != 'POST':
        return redirect('lista_usuarios')
 
    usuario = get_object_or_404(User, id=usuario_id)
 
    # Protección: no se puede desactivar a otro admin ni a uno mismo
    if usuario.rol == 'admin':
        messages.error(request, 'No se puede desactivar la cuenta de un administrador.')
        return redirect('lista_usuarios')
 
    if usuario == request.user:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('lista_usuarios')
 
    accion = request.POST.get('accion')
 
    if accion == 'desactivar':
        usuario.is_active = False
        usuario.save()
        # Cerrar todas las sesiones activas del usuario
        sesiones = Session.objects.filter(expire_date__gte=timezone.now())
        for sesion in sesiones:
            datos = sesion.get_decoded()
            if datos.get('_auth_user_id') == str(usuario.id):
                sesion.delete()
        messages.success(request, f'La cuenta de {usuario.get_full_name() or usuario.username} ha sido desactivada.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Cuenta de "{usuario.username}" desactivada — sesiones cerradas forzosamente', 'WARNING')

    elif accion == 'activar':
        usuario.is_active = True
        usuario.save()
        messages.success(request, f'La cuenta de {usuario.get_full_name() or usuario.username} ha sido reactivada.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Cuenta de "{usuario.username}" (ID {usuario.id}) reactivada', 'INFO')
 
    return redirect('lista_usuarios')
 
 
@login_required
@no_cache_view
def lista_medicos(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    from medicos.models import Medico
    medicos = Medico.objects.all()
    return render(request, 'admin/lista_medicos.html', {
        'medicos': medicos,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
@login_required
@no_cache_view
def lista_pacientes(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    pacientes_prenatales = Paciente.objects.filter(
        q_pacientes_con_acceso_prenatal()
    ).select_related('usuario').distinct()
    pacientes_generales = Paciente.objects.exclude(
        q_pacientes_con_acceso_prenatal()
    ).select_related('usuario').distinct()

    return render(request, 'admin/lista_pacientes.html', {
        'pacientes_prenatales': pacientes_prenatales,
        'pacientes_generales': pacientes_generales,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
@login_required
@no_cache_view
def todas_citas(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    # ── AUTO-CANCELAR citas pendientes (30 min de margen) ──
    auto_cancelar_citas(Cita.objects)
    hoy = timezone.now().date()

    citas_prenatales = Cita.objects.filter(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').distinct().order_by('-fecha', 'hora')

    citas_generales = Cita.objects.exclude(
        medico__medico__especialidad__tipo='prenatal'
    ).select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')

    return render(request, 'admin/todas_citas.html', {
        'citas_prenatales': citas_prenatales,
        'citas_generales': citas_generales,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })
 
 
@login_required
@no_cache_view
def controles_admin(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
 
    controles = ControlPrenatal.objects.all()
    return render(request, 'admin/controles.html', {
        'controles': controles,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })




@login_required
def admin_crear_usuario(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        rol        = request.POST.get('rol', '')
        password   = request.POST.get('password', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_usuario.html', {
                'roles': ['admin','medico','secretaria','paciente'],
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name,
            email=email, rol=rol
        )

        # Crear perfiles según rol
        if rol == 'medico':
            from medicos.models import Medico
            Medico.objects.get_or_create(
                usuario=user,
                defaults={
                    'especialidad': request.POST.get('especialidad', ''),
                    'telefono': request.POST.get('telefono_med', '')
                }
            )
        elif rol == 'paciente':
            from pacientes.models import Paciente
            Paciente.objects.get_or_create(usuario=user)

        messages.success(request, f'Usuario "{username}" creado correctamente.')
        registrar_log(request, 'CREATE', 'Usuarios',
            f'Usuario "{username}" creado con rol "{rol}"', 'INFO')
        return redirect('lista_usuarios')

    return render(request, 'admin/crear_usuario.html', {
        'roles': ['admin', 'medico', 'secretaria', 'paciente'],
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_editar_usuario(request, usuario_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    usuario = get_object_or_404(User, id=usuario_id)

    if request.method == 'POST':
        usuario.first_name = request.POST.get('first_name', '').strip()
        usuario.last_name  = request.POST.get('last_name', '').strip()
        usuario.email      = request.POST.get('email', '').strip()
        nuevo_rol = request.POST.get('rol', usuario.rol)
        usuario.rol = nuevo_rol
        password = request.POST.get('password', '').strip()
        if password:
            usuario.set_password(password)
        usuario.save()

        # Actualizar perfil médico si aplica
        if nuevo_rol == 'medico':
            from medicos.models import Medico
            medico, _ = Medico.objects.get_or_create(usuario=usuario)
            medico.especialidad = request.POST.get('especialidad', medico.especialidad)
            medico.telefono     = request.POST.get('telefono_med', medico.telefono)
            medico.save()

        messages.success(request, f'Usuario "{usuario.username}" actualizado correctamente.')
        registrar_log(request, 'UPDATE', 'Usuarios',
            f'Usuario "{usuario.username}" (ID {usuario.id}) actualizado — rol: {nuevo_rol}', 'INFO')
        return redirect('lista_usuarios')

    medico_perfil = None
    if usuario.rol == 'medico':
        from medicos.models import Medico
        try:
            medico_perfil = Medico.objects.get(usuario=usuario)
        except Medico.DoesNotExist:
            pass

    return render(request, 'admin/editar_usuario.html', {
        'usuario': usuario,
        'roles': ['admin', 'medico', 'secretaria', 'paciente'],
        'medico_perfil': medico_perfil,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_usuario(request, usuario_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    usuario = get_object_or_404(User, id=usuario_id)

    if usuario == request.user:
        messages.error(request, 'No puedes eliminar tu propia cuenta.')
        return redirect('lista_usuarios')

    if request.method == 'POST':
        nombre = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{nombre}" eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Usuarios',
            f'Usuario "{nombre}" (ID {usuario_id}) eliminado del sistema', 'WARNING')
        return redirect('lista_usuarios')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': usuario.get_full_name() or usuario.username,
        'tipo': 'usuario',
        'volver': 'lista_usuarios',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── PACIENTES ─────────────────────────────────

@login_required
def admin_crear_paciente(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')

        if User.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_paciente.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        user = User.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name,
            email=email, rol='paciente'
        )
        from pacientes.models import Paciente
        paciente, _ = Paciente.objects.get_or_create(usuario=user)
        paciente.cedula    = request.POST.get('cedula', '')
        paciente.telefono  = request.POST.get('telefono', '')
        paciente.edad      = request.POST.get('edad') or None
        paciente.direccion = request.POST.get('direccion', '')
        fecha_um = request.POST.get('fecha_ultima_menstruacion')
        fecha_pp = request.POST.get('fecha_probable_parto')
        paciente.fecha_ultima_menstruacion = fecha_um if fecha_um else None
        paciente.fecha_probable_parto      = fecha_pp if fecha_pp else None
        paciente.save()

        messages.success(request, f'Paciente "{first_name} {last_name}" creada correctamente.')
        registrar_log(request, 'CREATE', 'Pacientes',
            f'Paciente "{first_name} {last_name}" (usuario: {username}) registrada', 'INFO')
        return redirect('lista_pacientes')

    return render(request, 'admin/crear_paciente.html', {
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_editar_paciente(request, paciente_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if request.method == 'POST':
        paciente.usuario.first_name = request.POST.get('first_name', '').strip()
        paciente.usuario.last_name  = request.POST.get('last_name', '').strip()
        paciente.usuario.email      = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        if password:
            paciente.usuario.set_password(password)
        paciente.usuario.save()

        paciente.cedula    = request.POST.get('cedula', '')
        paciente.telefono  = request.POST.get('telefono', '')
        paciente.edad      = request.POST.get('edad') or None
        paciente.direccion = request.POST.get('direccion', '')
        fecha_um = request.POST.get('fecha_ultima_menstruacion')
        fecha_pp = request.POST.get('fecha_probable_parto')
        paciente.fecha_ultima_menstruacion = fecha_um if fecha_um else None
        paciente.fecha_probable_parto      = fecha_pp if fecha_pp else None
        paciente.save()

        messages.success(request, 'Paciente actualizada correctamente.')
        registrar_log(request, 'UPDATE', 'Pacientes',
            f'Datos de paciente "{paciente.usuario.get_full_name()}" (ID {paciente.id}) actualizados', 'INFO')
        return redirect('lista_pacientes')

    return render(request, 'admin/editar_paciente.html', {
        'paciente': paciente,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_paciente(request, paciente_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from pacientes.models import Paciente
    paciente = get_object_or_404(Paciente, id=paciente_id)

    if request.method == 'POST':
        nombre = str(paciente)
        paciente.usuario.delete()
        messages.success(request, f'Paciente "{nombre}" eliminada correctamente.')
        registrar_log(request, 'DELETE', 'Pacientes',
            f'Paciente "{nombre}" (ID {paciente_id}) eliminada del sistema', 'WARNING')
        return redirect('lista_pacientes')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'{paciente.usuario.first_name} {paciente.usuario.last_name}',
        'tipo': 'paciente',
        'volver': 'lista_pacientes',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── MÉDICOS ───────────────────────────────────
@login_required
def admin_crear_medico(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    from landing.models import Especialidad

    especialidades = Especialidad.objects.filter(activo=True)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')
        telefono   = request.POST.get('telefono', '').strip()
        especialidad_id = request.POST.get('especialidad_id', '').strip()

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_medico.html', {
                'form_data': request.POST,
                'especialidades': especialidades,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        user = Usuario.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name,
            email=email, rol='medico'
        )

        especialidad_obj = None
        if especialidad_id:
            try:
                especialidad_obj = Especialidad.objects.get(id=especialidad_id)
            except Especialidad.DoesNotExist:
                pass

        Medico.objects.create(
            usuario=user,
            especialidad=especialidad_obj,
            telefono=telefono
        )
        messages.success(request, f'Médico "{first_name} {last_name}" creado correctamente.')
        registrar_log(request, 'CREATE', 'Médicos',
            f'Médico "{first_name} {last_name}" (usuario: {username}) registrado', 'INFO')
        return redirect('lista_medicos')

    return render(request, 'admin/crear_medico.html', {
        'especialidades': especialidades,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_editar_medico(request, medico_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    from landing.models import Especialidad

    medico = get_object_or_404(Medico, id=medico_id)
    especialidades = Especialidad.objects.filter(activo=True)

    if request.method == 'POST':
        medico.usuario.first_name = request.POST.get('first_name', '').strip()
        medico.usuario.last_name  = request.POST.get('last_name', '').strip()
        medico.usuario.email      = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        if password:
            medico.usuario.set_password(password)
        medico.usuario.save()

        especialidad_id = request.POST.get('especialidad_id', '').strip()
        if especialidad_id:
            try:
                medico.especialidad = Especialidad.objects.get(id=especialidad_id)
            except Especialidad.DoesNotExist:
                medico.especialidad = None
        else:
            medico.especialidad = None

        medico.telefono = request.POST.get('telefono', '').strip()
        medico.save()

        messages.success(request, 'Médico actualizado correctamente.')
        registrar_log(request, 'UPDATE', 'Médicos',
            f'Datos del médico "{medico.usuario.get_full_name()}" (ID {medico.id}) actualizados', 'INFO')
        return redirect('lista_medicos')

    return render(request, 'admin/editar_medico.html', {
        'medico': medico,
        'especialidades': especialidades,
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


@login_required
def admin_eliminar_medico(request, medico_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    from medicos.models import Medico
    medico = get_object_or_404(Medico, id=medico_id)

    if request.method == 'POST':
        nombre = str(medico)
        medico.usuario.delete()
        messages.success(request, f'Médico "{nombre}" eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Médicos',
            f'Médico "{nombre}" (ID {medico_id}) eliminado del sistema', 'WARNING')
        return redirect('lista_medicos')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Dr. {medico.usuario.first_name} {medico.usuario.last_name}',
        'tipo': 'médico',
        'volver': 'lista_medicos',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── CITAS ─────────────────────────────────────

@login_required
def admin_eliminar_cita(request, cita_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    cita = get_object_or_404(Cita, id=cita_id)

    if request.method == 'POST':
        cita.delete()
        messages.success(request, 'Cita eliminada correctamente.')
        registrar_log(request, 'DELETE', 'Citas',
            f'Cita ID {cita_id} eliminada', 'WARNING')
        return redirect('todas_citas')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Cita de {cita.paciente.get_full_name()} — {cita.fecha}',
        'tipo': 'cita',
        'volver': 'todas_citas',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })


# ── CONTROLES ─────────────────────────────────

@login_required
def admin_eliminar_control(request, control_id):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    control = get_object_or_404(ControlPrenatal, id=control_id)

    if request.method == 'POST':
        control.delete()
        messages.success(request, 'Control prenatal eliminado correctamente.')
        registrar_log(request, 'DELETE', 'Controles Prenatales',
            f'Control prenatal ID {control_id} eliminado', 'WARNING')
        return redirect('controles_admin')

    return render(request, 'admin/confirmar_eliminar.html', {
        'objeto': f'Control de {control.paciente.get_full_name()} — {control.fecha}',
        'tipo': 'control prenatal',
        'volver': 'controles_admin',
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })

@login_required
@no_cache_view
def admin_crear_paciente_general(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        password   = request.POST.get('password', '')
        cedula     = request.POST.get('cedula', '').strip()
        telefono   = request.POST.get('telefono', '').strip()
        edad       = request.POST.get('edad') or None
        direccion  = request.POST.get('direccion', '').strip()

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, f'El usuario "{username}" ya existe.')
            return render(request, 'admin/crear_paciente_general.html', {
                'form_data': request.POST,
                'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
            })

        user = Usuario.objects.create_user(
            username=username, password=password,
            first_name=first_name, last_name=last_name,
            email=email, rol='paciente'
        )
        from pacientes.models import Paciente
        Paciente.objects.create(
            usuario=user, cedula=cedula, telefono=telefono,
            edad=edad, direccion=direccion
        )
        registrar_log(request, 'CREATE', 'Pacientes',
            f'Paciente general "{first_name} {last_name}" registrado', 'INFO')
        messages.success(request, f'Paciente general "{first_name} {last_name}" creado correctamente.')
        return redirect('lista_pacientes')

    return render(request, 'admin/crear_paciente_general.html', {
        'citas_pendientes': Cita.objects.filter(estado='pendiente').count(),
    })

@login_required
@login_required
@no_cache_view
def perfil_secretaria(request):
    if request.user.rol != 'secretaria':
        return redireccionar_por_rol(request.user)
 
    if request.method == 'POST':
        action = request.POST.get('action')
 
        if action == 'perfil':
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name  = request.POST.get('last_name', '')
            request.user.email      = request.POST.get('email', '')
            request.user.save()
            messages.success(request, 'Datos actualizados correctamente.')
 
        elif action == 'password':
            from django.contrib.auth import update_session_auth_hash
            password_actual    = request.POST.get('password_actual')
            password_nuevo     = request.POST.get('password_nuevo')
            password_confirmar = request.POST.get('password_confirmar')
 
            if not request.user.check_password(password_actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif password_nuevo != password_confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(password_nuevo) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(password_nuevo)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña actualizada correctamente.')
 
        return redirect('perfil_secretaria')
 
    return render(request, 'secretaria/perfil_secretaria.html')


# ═══════════════════════════════════════════════════════════════════════════
# CONSULTA GENERAL — Médico general registra consultas a pacientes generales
# ═══════════════════════════════════════════════════════════════════════════

@login_required
@no_cache_view
def registrar_consulta_general(request):
    """Registra una nueva consulta general — todos los médicos pueden usarla."""
    from paciente_general.models import ConsultaGeneral
    from citas.models import Cita
    from pacientes.models import Paciente

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    # Médico prenatal ve sus pacientes prenatales + pacientes con citas
    # Médico general ve solo sus pacientes con citas
    ids_citas = Cita.objects.filter(
        medico=request.user
    ).values_list('paciente_id', flat=True)

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except Exception:
        es_prenatal = False

    if es_prenatal:
        # Incluye pacientes con citas + pacientes que este médico tiene asignados como prenatal
        from django.db.models import Q
        pacientes_generales = Paciente.objects.filter(
            Q(usuario_id__in=ids_citas) | Q(medico_prenatal=request.user)
        ).select_related('usuario').distinct().order_by('usuario__first_name', 'usuario__last_name')
    else:
        pacientes_generales = Paciente.objects.filter(
            usuario_id__in=ids_citas
        ).select_related('usuario').order_by('usuario__first_name', 'usuario__last_name')

    # Pre-selección de paciente via ?pac=ID
    pac_preseleccionado = request.GET.get('pac') or request.POST.get('paciente')

    if request.method == 'POST':
        paciente_id = request.POST.get('paciente')
        if not paciente_id:
            messages.error(request, 'Debes seleccionar una paciente.')
            return render(request, 'medico/registrar_consulta_general.html',
                          {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})

        try:
            paciente_obj = Usuario.objects.get(id=paciente_id, rol='paciente')
        except Usuario.DoesNotExist:
            messages.error(request, 'Paciente no encontrada.')
            return render(request, 'medico/registrar_consulta_general.html',
                          {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})

        def _val(campo, default=''):
            v = request.POST.get(campo, default).strip()
            return v if v else default

        def _int(campo):
            try: return int(request.POST.get(campo, '')) or None
            except: return None

        def _float(campo):
            try: return float(request.POST.get(campo, '')) or None
            except: return None

        consulta = ConsultaGeneral.objects.create(
            paciente                 = paciente_obj,
            medico                   = request.user,
            motivo_consulta          = _val('motivo_consulta', '(sin motivo)'),
            antecedentes_personales  = _val('antecedentes_personales'),
            antecedentes_familiares  = _val('antecedentes_familiares'),
            antecedentes_alergicos   = _val('antecedentes_alergicos'),
            antecedentes_quirurgicos = _val('antecedentes_quirurgicos'),
            antecedentes_obstetricos = _val('antecedentes_obstetricos'),
            examen_fisico            = _val('examen_fisico'),
            presion_arterial         = _val('presion_arterial'),
            saturacion_oxigeno       = _int('saturacion_oxigeno'),
            frecuencia_cardiaca      = _int('frecuencia_cardiaca'),
            frecuencia_respiratoria  = _int('frecuencia_respiratoria'),
            temperatura              = _float('temperatura'),
            talla                    = _float('talla'),
            peso                     = _float('peso'),
            examenes_enviados        = _val('examenes_enviados'),
            evolucion_enfermedad     = _val('evolucion_enfermedad'),
            plan                     = _val('plan'),
            tratamiento              = _val('tratamiento'),
            diagnostico_1_patologia  = _val('diag1_patologia'),
            diagnostico_1_cie10      = _val('diag1_cie10'),
            diagnostico_1_presuntivo = bool(request.POST.get('diag1_presuntivo')),
            diagnostico_1_definitivo = bool(request.POST.get('diag1_definitivo')),
            diagnostico_2_patologia  = _val('diag2_patologia'),
            diagnostico_2_cie10      = _val('diag2_cie10'),
            diagnostico_2_presuntivo = bool(request.POST.get('diag2_presuntivo')),
            diagnostico_2_definitivo = bool(request.POST.get('diag2_definitivo')),
            diagnostico_3_patologia  = _val('diag3_patologia'),
            diagnostico_3_cie10      = _val('diag3_cie10'),
            diagnostico_3_presuntivo = bool(request.POST.get('diag3_presuntivo')),
            diagnostico_3_definitivo = bool(request.POST.get('diag3_definitivo')),
            proxima_cita             = request.POST.get('proxima_cita') or None,
        )
        messages.success(request, f'Consulta de {paciente_obj.get_full_name()} registrada correctamente.')
        return redirect('ver_consulta_general', consulta_id=consulta.id)

    return render(request, 'medico/registrar_consulta_general.html',
                  {'pacientes': pacientes_generales, 'pac_preseleccionado': pac_preseleccionado})


@login_required
@no_cache_view
def historial_consultas_generales(request):
    """Lista las consultas generales registradas por este médico específicamente."""
    from paciente_general.models import ConsultaGeneral

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    paciente_id = request.GET.get('paciente')

    # Solo consultas registradas por ESTE médico
    consultas = ConsultaGeneral.objects.filter(
        medico=request.user
    ).select_related('paciente', 'medico').order_by('-fecha')

    if paciente_id:
        consultas = consultas.filter(paciente_id=paciente_id)

    # Selector de pacientes: solo los que este médico ha atendido
    from citas.models import Cita
    ids_mis_pacientes = Cita.objects.filter(
        medico=request.user
    ).values_list('paciente_id', flat=True)

    from pacientes.models import Paciente
    pacientes_generales = Paciente.objects.filter(
        usuario_id__in=ids_mis_pacientes
    ).select_related('usuario').order_by('usuario__first_name')

    return render(request, 'medico/historial_consultas_generales.html', {
        'consultas': consultas,
        'pacientes': pacientes_generales,
        'paciente_filtrado_id': int(paciente_id) if paciente_id else None,
    })


@login_required
def ver_consulta_general(request, consulta_id):
    """Muestra el detalle completo de una consulta general."""
    from paciente_general.models import ConsultaGeneral

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    consulta = get_object_or_404(ConsultaGeneral, id=consulta_id)
    return render(request, 'medico/ver_consulta_general.html', {'consulta': consulta})


# ═══════════════════════════════════════════════════════════════════════════
# PROGRAMACIÓN DE PARTOS — Médico prenatal programa, paciente prenatal ve
# ═══════════════════════════════════════════════════════════════════════════

@login_required
@no_cache_view
def programar_parto(request):
    """El médico prenatal programa una fecha/hora de parto para una paciente."""
    from paciente_general.models import ProgramacionParto

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    pacientes_prenatales = Usuario.objects.filter(
        rol='paciente', is_active=True
    ).filter(q_usuarios_con_acceso_prenatal()).distinct().order_by('first_name', 'last_name')

    if request.method == 'POST':
        paciente_id      = request.POST.get('paciente')
        tipo             = request.POST.get('tipo', 'parto_natural')
        fecha_programada = request.POST.get('fecha_programada')
        hora_programada  = request.POST.get('hora_programada')
        semanas          = request.POST.get('semanas_gestacion') or None
        lugar            = request.POST.get('lugar', 'Zumedical — Centro Médico').strip()
        indicaciones     = request.POST.get('indicaciones', '').strip()

        if not all([paciente_id, fecha_programada, hora_programada]):
            messages.error(request, 'Paciente, fecha y hora son obligatorios.')
            return render(request, 'medico/programar_parto.html',
                          {'pacientes': pacientes_prenatales})

        try:
            paciente_obj = Usuario.objects.filter(
                id=paciente_id, rol='paciente'
            ).filter(q_usuarios_con_acceso_prenatal()).distinct().get()
        except Usuario.DoesNotExist:
            messages.error(request, 'Paciente prenatal no encontrada.')
            return render(request, 'medico/programar_parto.html',
                          {'pacientes': pacientes_prenatales})

        parto = ProgramacionParto.objects.create(
            paciente          = paciente_obj,
            medico            = request.user,
            tipo              = tipo,
            fecha_programada  = fecha_programada,
            hora_programada   = hora_programada,
            semanas_gestacion = int(semanas) if semanas else None,
            lugar             = lugar or 'Zumedical — Centro Médico',
            indicaciones      = indicaciones,
            estado            = 'programado',
        )
        messages.success(
            request,
            f'{parto.get_tipo_display()} programado para {paciente_obj.get_full_name()} '
            f'el {parto.fecha_programada.strftime("%d/%m/%Y")} a las {parto.hora_programada.strftime("%H:%M")}.'
        )
        return redirect('lista_programaciones_parto')

    return render(request, 'medico/programar_parto.html',
                  {'pacientes': pacientes_prenatales})


@login_required
@no_cache_view
def editar_parto(request, parto_id):
    """Edita o cambia el estado de una programación de parto."""
    from paciente_general.models import ProgramacionParto

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    parto = get_object_or_404(ProgramacionParto, id=parto_id)

    if request.method == 'POST':
        parto.tipo             = request.POST.get('tipo', parto.tipo)
        parto.fecha_programada = request.POST.get('fecha_programada', str(parto.fecha_programada))
        parto.hora_programada  = request.POST.get('hora_programada', str(parto.hora_programada))
        parto.lugar            = request.POST.get('lugar', parto.lugar).strip() or parto.lugar
        parto.indicaciones     = request.POST.get('indicaciones', parto.indicaciones).strip()
        parto.estado           = request.POST.get('estado', parto.estado)
        sem = request.POST.get('semanas_gestacion')
        if sem:
            try: parto.semanas_gestacion = int(sem)
            except: pass
        parto.save()
        messages.success(request, 'Programación de parto actualizada.')
        return redirect('lista_programaciones_parto')

    pacientes_prenatales = Usuario.objects.filter(
        rol='paciente', is_active=True
    ).filter(q_usuarios_con_acceso_prenatal()).distinct().order_by('first_name')
    return render(request, 'medico/programar_parto.html', {
        'pacientes': pacientes_prenatales,
        'parto':     parto,
        'modo':      'editar',
    })


@login_required
@no_cache_view
def lista_programaciones_parto(request):
    """Lista todas las programaciones de parto del médico."""
    from paciente_general.models import ProgramacionParto

    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    programaciones = ProgramacionParto.objects.select_related(
        'paciente', 'medico'
    ).order_by('fecha_programada')

    return render(request, 'medico/lista_programaciones_parto.html',
                  {'programaciones': programaciones})


@login_required
@no_cache_view
def perfil_medico(request):
    if request.user.rol != 'medico':
        return redireccionar_por_rol(request.user)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'perfil':
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name  = request.POST.get('last_name', '')
            request.user.email      = request.POST.get('email', '')
            request.user.save()
            messages.success(request, 'Datos actualizados correctamente.')

        elif action == 'password':
            actual    = request.POST.get('password_actual')
            nueva     = request.POST.get('password_nuevo')
            confirmar = request.POST.get('password_confirmar')

            if not request.user.check_password(actual):
                messages.error(request, 'La contraseña actual es incorrecta.')
            elif nueva != confirmar:
                messages.error(request, 'Las contraseñas nuevas no coinciden.')
            elif len(nueva) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            else:
                request.user.set_password(nueva)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, 'Contraseña actualizada correctamente.')

        return redirect('perfil_medico')

    try:
        es_prenatal = request.user.medico.especialidad and \
                      request.user.medico.especialidad.tipo == 'prenatal'
    except:
        es_prenatal = True

    if es_prenatal:
        return render(request, 'medico/perfil_medico.html', {'es_prenatal': es_prenatal})
    else:
        return render(request, 'medico/perfil_medico_general.html', {'es_prenatal': es_prenatal})


# ── AUDITORÍA ──────────────────────────────────────────────────

@login_required
@no_cache_view
def auditoria_admin(request):
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)

    hoy = timezone.now().date()

    # Queryset base
    logs_qs = LogAuditoria.objects.select_related('usuario').all()

    # Filtros GET
    accion    = request.GET.get('accion', '')
    severidad = request.GET.get('severidad', '')
    usuario   = request.GET.get('usuario', '')
    desde     = request.GET.get('desde', '')
    hasta     = request.GET.get('hasta', '')

    if accion:
        logs_qs = logs_qs.filter(accion=accion)
    if severidad:
        logs_qs = logs_qs.filter(severidad=severidad)
    if usuario:
        logs_qs = logs_qs.filter(usuario__id=usuario)
    if desde:
        logs_qs = logs_qs.filter(fecha__date__gte=desde)
    if hasta:
        logs_qs = logs_qs.filter(fecha__date__lte=hasta)

    # Exportar Excel
    if request.GET.get('export') == 'csv':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse(
                'El módulo openpyxl no está instalado. Ejecuta: pip install openpyxl',
                status=500, content_type='text/plain'
            )
        from django.utils import timezone as tz
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Auditoría Zumedical"

        # Título principal
        ws.merge_cells('A1:H1')
        titulo = ws['A1']
        titulo.value = f'Reporte de Auditoría — Zumedical — {hoy.strftime("%d/%m/%Y")}'
        titulo.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
        titulo.fill = PatternFill('solid', fgColor='8A2563')
        titulo.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Encabezados
        headers = ['N°', 'Fecha', 'Hora', 'Usuario', 'Rol', 'Acción', 'Módulo', 'IP', 'Severidad', 'Descripción']
        ws.append([])  # fila vacía
        ws.append(headers)
        header_row = ws.max_row
        header_fill = PatternFill('solid', fgColor='CC4D99')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        thin = Side(style='thin', color='E8AAD4')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_num, _ in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[header_row].height = 22

        # Datos
        fill_par = PatternFill('solid', fgColor='FBF0F7')
        fill_imp = PatternFill('solid', fgColor='FFFFFF')
        font_data = Font(name='Calibri', size=10)

        for i, log in enumerate(logs_qs.order_by('-fecha'), 1):
            fecha_local = tz.localtime(log.fecha)
            nombre = log.usuario.get_full_name() if log.usuario else '—'
            rol = log.usuario.rol.capitalize() if log.usuario else '—'
            row = [
                i,
                fecha_local.strftime('%d/%m/%Y'),
                fecha_local.strftime('%H:%M:%S'),
                nombre,
                rol,
                log.get_accion_display(),
                log.modulo or '—',
                log.ip_address or '—',
                log.get_severidad_display(),
                log.descripcion or '—',
            ]
            ws.append(row)
            current_row = ws.max_row
            fill = fill_par if i % 2 == 0 else fill_imp
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = font_data
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=col_num == len(row))

        # Anchos de columna
        anchos = [6, 12, 10, 22, 12, 16, 16, 14, 12, 50]
        for col_num, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = ancho

        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="auditoria_zumedical_{hoy}.xlsx"'
        return response

    # KPIs — calcular rango UTC del día local en Ecuador
    from django.utils import timezone as tz
    from zoneinfo import ZoneInfo
    from datetime import datetime, timedelta

    ecuador = ZoneInfo('America/Guayaquil')
    ahora_ecuador = tz.now().astimezone(ecuador)
    hoy_local = ahora_ecuador.date()

    # Convertir inicio y fin del día Ecuador a UTC para filtrar correctamente
    inicio_dia_ec = datetime(hoy_local.year, hoy_local.month, hoy_local.day, 0, 0, 0, tzinfo=ecuador)
    fin_dia_ec    = datetime(hoy_local.year, hoy_local.month, hoy_local.day, 23, 59, 59, tzinfo=ecuador)
    inicio_utc = inicio_dia_ec.astimezone(ZoneInfo('UTC'))
    fin_utc    = fin_dia_ec.astimezone(ZoneInfo('UTC'))

    total_logs           = LogAuditoria.objects.count()
    logs_criticos_hoy    = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc, severidad='CRITICAL').count()
    logins_hoy           = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc, accion='LOGIN').count()
    usuarios_activos_hoy = LogAuditoria.objects.filter(
        fecha__gte=inicio_utc, fecha__lte=fin_utc).values('usuario').distinct().count()
    
    # Paginación: 25 registros por página
    paginator = Paginator(logs_qs, 25)
    page_num  = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_num)

    return render(request, 'admin/auditoria_admin.html', {
        'logs':                  logs_page,
        'total_logs':            total_logs,
        'logs_criticos_hoy':     logs_criticos_hoy,
        'logins_hoy':            logins_hoy,
        'usuarios_activos_hoy':  usuarios_activos_hoy,
        'usuarios_lista':        User.objects.filter(is_active=True).order_by('first_name'),
        'citas_pendientes':      Cita.objects.filter(estado='pendiente').count(),
        'config_retencion':      request.session.get('auditoria_retencion', '90'),
    })


@login_required
@no_cache_view
def auditoria_admin_data(request):
    """
    Endpoint JSON para el panel de Auditoria.
    Devuelve KPIs, datos para graficos (actividad 7 dias + dona por accion)
    y filas de tabla paginadas, respetando los mismos filtros GET que la
    vista principal. Usado para busqueda en tiempo real, filtros sin
    recargar, auto-refresh y los graficos.
    """
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    hoy = timezone.now().date()

    logs_qs = LogAuditoria.objects.select_related('usuario').all()

    accion    = request.GET.get('accion', '')
    severidad = request.GET.get('severidad', '')
    usuario   = request.GET.get('usuario', '')
    desde     = request.GET.get('desde', '')
    hasta     = request.GET.get('hasta', '')
    q         = request.GET.get('q', '').strip()

    if accion:
        logs_qs = logs_qs.filter(accion=accion)
    if severidad:
        logs_qs = logs_qs.filter(severidad=severidad)
    if usuario:
        logs_qs = logs_qs.filter(usuario__id=usuario)
    if desde:
        logs_qs = logs_qs.filter(fecha__date__gte=desde)
    if hasta:
        logs_qs = logs_qs.filter(fecha__date__lte=hasta)
    if q:
        logs_qs = logs_qs.filter(
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q) |
            Q(usuario__username__icontains=q) |
            Q(modulo__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(ip_address__icontains=q)
        )

    from zoneinfo import ZoneInfo
    from datetime import datetime
    ecuador = ZoneInfo('America/Guayaquil')
    ahora_ec = timezone.now().astimezone(ecuador)
    hoy_ec = ahora_ec.date()

    # Rango UTC del día local en Ecuador
    inicio_dia_ec = datetime(hoy_ec.year, hoy_ec.month, hoy_ec.day, 0, 0, 0, tzinfo=ecuador)
    fin_dia_ec    = datetime(hoy_ec.year, hoy_ec.month, hoy_ec.day, 23, 59, 59, tzinfo=ecuador)
    inicio_utc = inicio_dia_ec.astimezone(ZoneInfo('UTC'))
    fin_utc    = fin_dia_ec.astimezone(ZoneInfo('UTC'))

    total_logs           = LogAuditoria.objects.count()
    logs_criticos_hoy    = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc, severidad='CRITICAL').count()
    logins_hoy           = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc, accion='LOGIN').count()
    usuarios_activos_hoy = LogAuditoria.objects.filter(fecha__gte=inicio_utc, fecha__lte=fin_utc).values('usuario').distinct().count()

    # Grafico 1: actividad por dia (ultimos 7 dias) — usando fecha Ecuador
    actividad_labels = []
    actividad_data = []
    utc = ZoneInfo('UTC')
    for i in range(6, -1, -1):
        dia = hoy_ec - timedelta(days=i)
        # Convertir el día Ecuador a rango UTC para filtrar correctamente
        inicio = datetime(dia.year, dia.month, dia.day, 0, 0, 0, tzinfo=ecuador).astimezone(utc)
        fin    = datetime(dia.year, dia.month, dia.day, 23, 59, 59, tzinfo=ecuador).astimezone(utc)
        count = LogAuditoria.objects.filter(fecha__range=(inicio, fin)).count()
        actividad_labels.append(dia.strftime('%d/%m'))
        actividad_data.append(count)

    # Grafico 2: dona por tipo de accion (sobre el queryset filtrado)
    acciones_count = (
        logs_qs.values('accion')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    try:
        accion_labels_map = dict(LogAuditoria._meta.get_field('accion').choices)
    except Exception:
        accion_labels_map = {}
    dona_labels = [accion_labels_map.get(a['accion'], a['accion']) for a in acciones_count]
    dona_data   = [a['total'] for a in acciones_count]

    paginator = Paginator(logs_qs, 25)
    page_num  = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_num)

    rows = []
    for log in logs_page:
        rows.append({
            'id':          log.id,
            'fecha':       log.fecha.strftime('%d/%m/%Y'),
            'hora':        log.fecha.strftime('%H:%M:%S'),
            'usuario':     log.usuario.get_full_name() if log.usuario and log.usuario.get_full_name() else (log.usuario.username if log.usuario else '\u2014'),
            'usuario_inicial': (log.usuario.first_name[:1] if log.usuario and log.usuario.first_name else (log.usuario.username[:1] if log.usuario else '?')).upper(),
            'rol':         log.usuario.groups.first().name if log.usuario and log.usuario.groups.first() else 'Usuario',
            'accion':      log.accion,
            'accion_display': log.get_accion_display(),
            'modulo':      log.modulo or '\u2014',
            'ip':          log.ip_address or '\u2014',
            'severidad':   log.severidad,
            'severidad_display': log.get_severidad_display(),
            'descripcion': log.descripcion or 'Sin descripcion',
        })

    return JsonResponse({
        'kpis': {
            'total_logs': total_logs,
            'logs_criticos_hoy': logs_criticos_hoy,
            'logins_hoy': logins_hoy,
            'usuarios_activos_hoy': usuarios_activos_hoy,
        },
        'charts': {
            'actividad_labels': actividad_labels,
            'actividad_data': actividad_data,
            'dona_labels': dona_labels,
            'dona_data': dona_data,
        },
        'rows': rows,
        'pagination': {
            'current_page': logs_page.number,
            'num_pages': paginator.num_pages,
            'count': paginator.count,
            'has_previous': logs_page.has_previous(),
            'has_next': logs_page.has_next(),
        },
    })
@login_required
@no_cache_view
def auditoria_config(request):
    """Guarda la configuración de retención en sesión."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
    if request.method == 'POST':
        dias = request.POST.get('retencion', '90')
        if dias in ['0', '30', '90', '180', '365']:
            request.session['auditoria_retencion'] = dias
            messages.success(request, f'Configuración guardada: {"Nunca eliminar" if dias == "0" else f"{dias} días de retención"}.')
        else:
            messages.error(request, 'Valor de retención inválido.')
    return redirect('auditoria_admin')
 
 
@login_required
@no_cache_view
def auditoria_limpiar(request):
    """Elimina registros de auditoría según la retención configurada."""
    if request.user.rol != 'admin':
        return redireccionar_por_rol(request.user)
    if request.method == 'POST':
        dias = request.session.get('auditoria_retencion', '90')
        if dias == '0':
            messages.warning(request, 'La configuración actual es "Nunca eliminar". No se eliminó ningún registro.')
        else:
            from django.utils import timezone
            from datetime import timedelta
            limite = timezone.now() - timedelta(days=int(dias))
            eliminados = LogAuditoria.objects.filter(fecha__lt=limite).count()
            LogAuditoria.objects.filter(fecha__lt=limite).delete()
            registrar_log(request, 'DELETE', 'Auditoría',
                          f'Eliminados {eliminados} registros con más de {dias} días.', 'WARNING')
            messages.success(request, f'Se eliminaron {eliminados} registros con más de {dias} días.')
    return redirect('auditoria_admin')

# ── ENDPOINTS DE REPORTES ─────────────────────────────────────

@login_required
def reporte_pacientes_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from pacientes.models import Paciente
    pacientes = Paciente.objects.select_related('usuario').order_by('usuario__last_name')
    rows = []
    for i, p in enumerate(pacientes, 1):
        rows.append({
            'id':         i,
            'codigo':     f'PAC-{p.id:04d}',
            'cedula':     p.cedula or '—',
            'nombre':     p.usuario.get_full_name() or p.usuario.username,
            'edad':       p.edad if p.edad else '—',
            'telefono':   p.telefono or '—',
            'email':      p.usuario.email or '—',
            'direccion':  p.direccion or '—',
            'registro':   p.usuario.date_joined.strftime('%d/%m/%Y'),
            'estado':     'Activa' if p.usuario.is_active else 'Inactiva',
            'tipo':       'Prenatal' if p.usuario.puede_prenatal else 'General',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_medicos_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from medicos.models import Medico
    medicos = Medico.objects.select_related('usuario', 'especialidad').order_by('usuario__last_name')
    rows = []
    for i, m in enumerate(medicos, 1):
        rows.append({
            'id':           f'MED-{m.id:04d}',
            'nombre':       m.usuario.get_full_name() or m.usuario.username,
            'especialidad': m.especialidad.nombre if m.especialidad else '—',
            'telefono':     m.telefono or '—',
            'email':        m.usuario.email or '—',
            'estado':       'Activo' if m.usuario.is_active else 'Inactivo',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_citas_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from citas.models import Cita
    citas = Cita.objects.select_related('paciente', 'medico', 'especialidad').order_by('-fecha', 'hora')
    rows = []
    for i, c in enumerate(citas, 1):
        rows.append({
            'numero':       i,
            'paciente':     c.paciente.get_full_name() or c.paciente.username,
            'medico':       c.medico.get_full_name() or c.medico.username,
            'especialidad': c.especialidad.nombre if c.especialidad else '—',
            'fecha':        c.fecha.strftime('%d/%m/%Y'),
            'hora':         c.hora.strftime('%H:%M'),
            'estado':       c.get_estado_display(),
            'motivo':       c.motivo or '—',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_controles_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from control_prenatal.models import ControlPrenatal
    controles = ControlPrenatal.objects.select_related('paciente', 'medico').order_by('-fecha')
    rows = []
    for i, c in enumerate(controles, 1):
        rows.append({
            'numero':        i,
            'paciente':      c.paciente.get_full_name() or c.paciente.username,
            'medico':        c.medico.get_full_name() or c.medico.username,
            'fecha':         c.fecha.strftime('%d/%m/%Y'),
            'semanas':       c.semanas_gestacion,
            'peso':          f'{c.peso} kg',
            'presion':       c.presion_arterial or '—',
            'observaciones': c.observaciones or '—',
        })
    return JsonResponse({'rows': rows})


@login_required
def reporte_usuarios_data(request):
    if request.user.rol != 'admin':
        return JsonResponse({'error': 'No autorizado'}, status=403)

    from django.contrib.sessions.models import Session
    from django.utils import timezone as tz

    sesiones_activas = Session.objects.filter(expire_date__gte=tz.now())
    ids_activos = set()
    for s in sesiones_activas:
        uid = s.get_decoded().get('_auth_user_id')
        if uid:
            ids_activos.add(int(uid))

    usuarios = User.objects.exclude(rol='paciente').order_by('rol', 'last_name')
    rows = []
    for i, u in enumerate(usuarios, 1):
        rows.append({
            'id':            f'USR-{u.id:04d}',
            'usuario':       u.username,
            'nombre':        u.get_full_name() or '—',
            'rol':           u.rol.capitalize(),
            'email':         u.email or '—',
            'estado':        'Activo' if u.is_active else 'Inactivo',
            'ultimo_acceso': u.last_login.strftime('%d/%m/%Y %H:%M') if u.last_login else 'Nunca',
            'en_linea':      u.id in ids_activos,
        })
    return JsonResponse({'rows': rows})
